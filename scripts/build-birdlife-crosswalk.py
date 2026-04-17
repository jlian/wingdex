#!/usr/bin/env python3
"""Build a BirdLife -> eBird crosswalk from AviList and export as JSON."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl first")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
AVILIST_CANDIDATES = [
    ROOT / ".tmp" / "AviList-v2025-11Jun-extended.xlsx",
    ROOT / ".tmp" / "avilist-v2025-extended.xlsx",
]
AVILIST_PATH = next((p for p in AVILIST_CANDIDATES if p.exists()), AVILIST_CANDIDATES[0])
if not AVILIST_PATH.exists():
    print("AviList workbook not found. Expected one of:", file=sys.stderr)
    for candidate in AVILIST_CANDIDATES:
        print(f"  - {candidate}", file=sys.stderr)
    sys.exit(1)
TAXONOMY_PATH = ROOT / "src" / "lib" / "taxonomy.json"
OUTPUT_PATH = ROOT / ".tmp" / "birdlife-shp" / "birdlife-crosswalk.json"

# Load eBird taxonomy
ebird = json.loads(TAXONOMY_PATH.read_text())
ebird_sci = {e[1].lower(): e[2] for e in ebird if len(e) > 2 and e[2]}
ebird_valid_codes = set(ebird_sci.values())  # set of valid eBird codes for fast lookup

# Parse AviList
wb = openpyxl.load_workbook(str(AVILIST_PATH), read_only=True)
ws = wb[wb.sheetnames[0]]

crosswalk = {}
# Per-eBird-code BirdLife DataZone species IDs (from AviList column 16).
# Species IDs are the numeric path segment of factsheet URLs like
# https://datazone.birdlife.org/species/factsheet/45020636
code_to_datazone = {}
# Also build a protonym index: maps old/original binomials to the PARENT
# species eBird code. BirdLife BOTW often uses protonyms as species names
# when they treat a subspecies as a full species (e.g., "Aethopyga latouchii"
# is the protonym of AviList subspecies "Aethopyga christinae latouchii",
# whose parent species code is fotsun1, not the subspecies code fotsun4).
protonym_to_code = {}
current_species_code = ""  # track the parent species code as we iterate
stats = {"total": 0, "with_code": 0, "direct": 0, "crosswalk": 0, "protonym_recovered": 0, "no_code": 0}

import re

def extract_bow_code(bow_url):
    """Extract species code from Birds of the World URL like .../bow/species/fotsun1/"""
    if not bow_url:
        return ""
    m = re.search(r'/bow/species/([a-z0-9]+)', str(bow_url))
    return m.group(1) if m else ""


def extract_datazone_id(url):
    """Extract numeric species ID from a DataZone factsheet URL like
    https://datazone.birdlife.org/species/factsheet/45020636"""
    if not url:
        return ""
    m = re.search(r'/species/factsheet/(\d+)', str(url))
    return m.group(1) if m else ""

for row in ws.iter_rows(min_row=2, values_only=True):
    stats["total"] += 1
    rank = (row[1] or "").strip()
    sci_name = (row[5] or "").strip()
    raw_code = (row[17] or "").strip()
    bow_url = (row[18] or "").strip()
    datazone_url = (row[16] or "").strip() if row[16] else ""
    protonym = (row[25] or "").strip()

    if not sci_name:
        continue

    # Prefer the BOW URL species code (more reliable than raw code column).
    # BOW URL only appears on species rows and always has the canonical code.
    bow_code = extract_bow_code(bow_url)

    # Track the current parent species code from species rows only.
    # Priority: BOW URL code > raw code (if valid) > our taxonomy lookup.
    # Validates raw code against taxonomy to catch AviList typos like fotswi1 -> fotswi.
    if rank == "species":
        validated_raw = raw_code if raw_code in ebird_valid_codes else ""
        current_species_code = bow_code or validated_raw or ebird_sci.get(sci_name.lower(), "") or raw_code
        datazone_id = extract_datazone_id(datazone_url)
        if datazone_id and current_species_code:
            code_to_datazone.setdefault(current_species_code, datazone_id)

    ebird_code = raw_code
    if not ebird_code:
        # Even without an AviList code, if the species is in our taxonomy,
        # we can still set current_species_code (done above) and skip this row.
        stats["no_code"] += 1
        if not current_species_code:
            continue
        # If we have a current_species_code from taxonomy lookup, keep processing
        # so protonyms still get indexed.
        ebird_code = current_species_code

    stats["with_code"] += 1
    sci_lower = sci_name.lower()

    # For protonym index, always use current_species_code (parent)
    if protonym and current_species_code:
        proto_lower = protonym.lower().strip()
        if proto_lower and proto_lower not in ebird_sci and proto_lower != sci_lower:
            protonym_to_code.setdefault(proto_lower, current_species_code)

    if sci_lower in ebird_sci:
        stats["direct"] += 1
    elif sci_lower not in crosswalk:
        # For subspecies rows, use the parent species code (not the subspp code)
        # For species rows, prefer BOW URL code (fixes cases like fotswi1 -> fotswi)
        if rank == "subspecies" and current_species_code:
            code_to_store = current_species_code
        elif rank == "species" and bow_code:
            code_to_store = bow_code
        else:
            code_to_store = current_species_code or ebird_code
        crosswalk[sci_lower] = code_to_store
        stats["crosswalk"] += 1

wb.close()

# Second pass: recover BirdLife names via protonym matching.
# BirdLife BOTW uses binomials like "Aethopyga latouchii" which are protonyms
# of AviList subspecies like "Aethopyga christinae latouchii".
try:
    import fiona
    gpkg = ROOT / ".tmp" / "birdlife-shp" / "BOTW_2025.gpkg"
    if gpkg.exists():
        birdlife_names = set()
        with fiona.open(str(gpkg), layer="all_species") as src:
            for feat in src:
                sci = (feat["properties"].get("sci_name") or "").strip()
                if sci:
                    birdlife_names.add(sci)

        for sci in sorted(birdlife_names):
            sci_lower = sci.lower()
            if sci_lower in ebird_sci or sci_lower in crosswalk:
                continue
            # Try protonym matching
            code = protonym_to_code.get(sci_lower)
            if code:
                crosswalk[sci_lower] = code
                stats["protonym_recovered"] += 1

        print(f"BirdLife unique species: {len(birdlife_names)}")
except ImportError:
    print("Note: fiona not available, skipping protonym recovery from GeoPackage")

wb.close()

print(f"AviList rows: {stats['total']}")
print(f"With eBird code: {stats['with_code']}")
print(f"Direct eBird match: {stats['direct']}")
print(f"Crosswalk (AviList name): {stats['crosswalk']}")
print(f"Protonym recovered: {stats['protonym_recovered']}")
print(f"No eBird code: {stats['no_code']}")
print(f"Protonym index size: {len(protonym_to_code)}")

# Verify our test case
for name in ["Aethopyga latouchii", "Aethopyga christinae"]:
    code = crosswalk.get(name.lower()) or ebird_sci.get(name.lower())
    print(f"  {name} -> {code}")

# Write crosswalk JSON (lowercase sci_name -> ebird_code)
OUTPUT_PATH.write_text(json.dumps(crosswalk, indent=2, sort_keys=True) + "\n")
print(f"\nWrote {len(crosswalk)} entries to {OUTPUT_PATH}")

# Hydrate taxonomy.json tuple slot [5] with BirdLife DataZone species IDs.
# Tuple layout: [common, scientific, ebirdCode, wikiTitle, thumbnailPath, birdlifeId]
print(f"\nBirdLife DataZone IDs collected: {len(code_to_datazone)}")
taxonomy = json.loads(TAXONOMY_PATH.read_text())
updated = 0
for entry in taxonomy:
    code = entry[2] if len(entry) > 2 else ""
    datazone_id = code_to_datazone.get(code) if code else None
    if not datazone_id:
        # Trim any stale slot-5 value without inflating rows that have none.
        if len(entry) >= 6:
            entry[5] = ""
        continue
    # Pad missing wikiTitle/thumbnail slots with empty strings so [5] lands
    # in the right column. The TS reader treats empty strings as absent.
    while len(entry) < 5:
        entry.append("")
    if len(entry) < 6:
        entry.append(datazone_id)
    else:
        entry[5] = datazone_id
    updated += 1

# Strip trailing empty strings per row to keep the file compact.
for entry in taxonomy:
    while len(entry) > 2 and entry[-1] == "":
        entry.pop()

TAXONOMY_PATH.write_text(
    "[\n"
    + ",\n".join(
        json.dumps(e, ensure_ascii=False, separators=(",", ":")) for e in taxonomy
    )
    + "\n]\n"
)
print(f"Updated {updated} taxonomy.json entries with BirdLife DataZone IDs")
