#!/usr/bin/env python3
"""Emit the EX/EW species list from the AviList workbook.

WHY THIS EXISTS RATHER THAN A HARDCODED LIST OF 143 ROW INDEXES
---------------------------------------------------------------
Row indexes are meaningless across taxonomy versions: dropping rows renumbers
everything after them, so a list of indexes is only valid for one exact file.
Deriving the exclusion from the IUCN status column means the next taxonomy
refresh re-derives it automatically instead of someone remembering to redo 143
deletions by hand.

The workbook is the same one scripts/build-birdlife-crosswalk.py already opens,
so this adds a column read rather than a new data source.

Usage:
  python3 scripts/emit-extinct-list.py \
      --avilist .tmp/AviList-v2025-11Jun-extended.xlsx \
      --taxonomy src/lib/taxonomy.json \
      --out scripts/extinct-species.json
"""
import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl first")

# IUCN codes that mean "cannot be photographed alive in the wild".
#
# Measured on AviList v2025b: EX 147, EW 5, CR (PE) 17, CR (PEW) 2, CR 183.
#
# EW is included because an Extinct in the Wild species exists only in
# captivity, so a wild-bird identifier should not offer it.
#
# CR (PE), "Critically Endangered, Possibly Extinct", is deliberately NOT
# excluded, and neither is plain CR. Ivory-billed Woodpecker sits in that group,
# and those are exactly the records that would matter most if one were ever
# photographed. Excluding them would make the app unable to report the single
# most significant sighting it could ever receive.
EXCLUDE = {"EX", "EW"}

# AviList v2025b column headers, matched case-insensitively. Falls back to a
# fuzzy search so a header rename does not silently exclude nothing.
SCI_HEADERS = ("scientific_name",)
STATUS_HEADERS = ("iucn_red_list_category",)
CODE_HEADERS = ("species_code_cornell_lab",)


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--avilist", required=True)
    ap.add_argument("--taxonomy", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    tax = json.loads(Path(args.taxonomy).read_text())
    by_sci = {}
    for i, row in enumerate(tax):
        if len(row) > 1 and row[1]:
            by_sci.setdefault(norm(row[1]), i)

    wb = openpyxl.load_workbook(args.avilist, read_only=True)
    ws = wb[wb.sheetnames[0]]

    by_code = {}
    for i, row in enumerate(tax):
        if len(row) > 2 and row[2]:
            by_code.setdefault(row[2], i)

    header = None
    sci_col = status_col = code_col = rank_col = None
    hits = []
    unmatched = []
    seen_status = {}

    def find(headers, exact, *fuzzy):
        for j, h in enumerate(headers):
            if h in exact:
                return j
        for j, h in enumerate(headers):
            if all(f in h for f in fuzzy):
                return j
        return None

    for row in ws.iter_rows(min_row=1, values_only=True):
        if header is None:
            header = [norm(c) for c in row]
            sci_col = find(header, SCI_HEADERS, "scientific")
            status_col = find(header, STATUS_HEADERS, "iucn")
            code_col = find(header, CODE_HEADERS, "species_code")
            rank_col = find(header, ("taxon_rank",), "rank")
            if sci_col is None or status_col is None:
                sys.exit("could not find scientific-name and IUCN columns; "
                         f"headers were: {header}")
            print(f"columns: scientific={sci_col} iucn={status_col} "
                  f"code={code_col} rank={rank_col}")
            continue

        # Subspecies rows carry their parent's status; only species rows map
        # onto our taxonomy, and counting both would inflate the total.
        if rank_col is not None and norm(row[rank_col]) != "species":
            continue

        sci = norm(row[sci_col])
        status = norm(row[status_col]).upper()
        if not sci or not status:
            continue
        seen_status[status] = seen_status.get(status, 0) + 1
        if status not in EXCLUDE:
            continue
        # Match on the eBird code first: it is the join key the rest of the
        # pipeline uses, and it survives a scientific-name revision.
        idx = None
        if code_col is not None and row[code_col]:
            idx = by_code.get(str(row[code_col]).strip())
        if idx is None:
            idx = by_sci.get(sci)
        if idx is None:
            unmatched.append((sci, status))
            continue
        hits.append(dict(idx=idx, scientific=tax[idx][1], common=tax[idx][0],
                         status=status))

    hits.sort(key=lambda h: h["idx"])
    print(f"IUCN status values seen: {dict(sorted(seen_status.items()))}")
    print(f"EX/EW rows matched into taxonomy: {len(hits)}")
    print(f"EX/EW rows NOT in our taxonomy:   {len(unmatched)}")

    out = dict(
        excluded_statuses=sorted(EXCLUDE),
        taxonomy_rows=len(tax),
        count=len(hits),
        species=[{k: h[k] for k in ("scientific", "common", "status")} for h in hits],
        indexes_for_reference_only=[h["idx"] for h in hits],
    )
    Path(args.out).write_text(json.dumps(out, indent=1) + "\n")
    print(f"wrote {args.out}")
    for h in hits[:10]:
        print(f"  {h['idx']:>6}  {h['status']}  {h['common']}")
    if len(hits) > 10:
        print(f"  ... and {len(hits)-10} more")


if __name__ == "__main__":
    main()
